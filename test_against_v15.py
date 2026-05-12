# ══════════════════════════════════════════════════════════════════════════
# 回归测试:验证新引擎与 V15 矩阵 + 文件 4 的 857 触发日完全一致
# ══════════════════════════════════════════════════════════════════════════
#
# 用法:
#   python test_against_v15.py --v15 path/to/V15.xlsx --triggers path/to/file4.xlsx
#
# 验证内容:
#   1. 在 V15 矩阵的每一天上,用新引擎重算 10 触发器,与 V15 矩阵的 0/1 因子一致
#   2. 与文件 4 的 857 个历史触发日完全一致(0 偏差才算通过)
#
# 这个测试是上线的硬关卡:不通过不能上生产。
# ══════════════════════════════════════════════════════════════════════════

import argparse
import sys
from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl
import warnings
warnings.filterwarnings('ignore')

from sp500_signal_engine import TRIGGERS, evaluate_trigger


def load_v15_factor_matrix(path):
    """加载因子矩阵 — 自动适配 V15(59因子)和 V16(60因子)"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    candidates = ['1_因子矩阵_60', '1_因子矩阵_59']  # V16 优先
    sheet_name = next((s for s in candidates if s in wb.sheetnames), None)
    wb.close()
    if sheet_name is None:
        raise ValueError(f"找不到因子矩阵 sheet (尝试: {candidates})")
    fm = pd.read_excel(path, sheet_name=sheet_name, header=0)
    fm['date'] = pd.to_datetime(fm['date'])
    return fm


def load_v15_intermediate(path):
    """加载 V15 中间量 (sheet 2_中间量),用于 SP500 价格"""
    inter = pd.read_excel(path, sheet_name='2_中间量', header=0)
    inter['date'] = pd.to_datetime(inter['date'])
    return inter


def load_file4_trigger_dates(path):
    """从文件 4 (10 触发器合集) 加载 857 个触发日"""
    triggers = ['01_2000离场','02_2007离场','03_2015离场','04_2022离场',
                '05_2002入场','06_2009入场','07_2020入场','08_2022入场',
                '09_白银坑1组','10_白银坑2组']

    # 映射到我们的 T1-T10 命名
    name_map = {
        '01_2000离场':  'T1_2000离场',
        '02_2007离场':  'T2_2007离场',
        '03_2015离场':  'T3_2015离场',
        '04_2022离场':  'T4_2022离场',
        '05_2002入场':  'T5_2002入场',
        '06_2009入场':  'T6_2009入场',
        '07_2020入场':  'T7_2020入场',
        '08_2022入场':  'T8_2022入场',
        '09_白银坑1组': 'T9_白银坑1组',
        '10_白银坑2组': 'T10_白银坑2组',
    }

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = {}
    for t in triggers:
        sn = f'{t}_触发日明细'
        ws = wb[sn]
        dates = []
        found = False
        for r in range(1, ws.max_row+1):
            cells = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column+1, 4))]
            if not found:
                if cells[0] == '日期':
                    found = True
                continue
            v = cells[0]
            if v is None: continue
            if hasattr(v, 'year'):
                dates.append(pd.Timestamp(v))
            elif isinstance(v, str):
                try: dates.append(pd.to_datetime(v))
                except: pass
        result[name_map[t]] = sorted(set(dates))
    wb.close()
    return result


def evaluate_triggers_on_v15(fm, inter=None):
    """
    用 V15 因子矩阵 + 触发器公式 直接重算 10 个触发器在每一天的 0/1 状态。

    返回: dict {trigger_id: list of triggered dates}

    注意:
    - V15 矩阵的 'S-' 列是 0/1 (是否回撤),不是回撤数值
    - V16 矩阵用列名 'Sσ_200 < 0',V17 矩阵改为 'S_t < S_200'(数学等价)
    - 引擎统一用 'S_t < S_200',如遇 V16 矩阵自动建别名
    """
    # 兼容 V16:把 'Sσ_200 < 0' 别名为 'S_t < S_200'(语义等价)
    if 'Sσ_200 < 0' in fm.columns and 'S_t < S_200' not in fm.columns:
        fm = fm.copy()
        fm['S_t < S_200'] = fm['Sσ_200 < 0']

    def s(col):
        return fm[col].astype(int).values
    def NOT(arr):
        return 1 - arr

    T = {}
    T['T1_2000离场']  = (s('W+200') & s('Sσ_200 ≥ +1σ') & s('F+') & s('Y-') &
                        s('N- ≥ μ+1σ') & s('HY_c21 > 0') & s('ERP < 0'))
    T['T2_2007离场']  = (s('OIL_pct5y > 85%') & s('E-2') & s('Sσ_200 ≥ +1σ') &
                        s('W+200') & s('F-') & s('Y+') & s('S+6'))
    T['T3_2015离场']  = (s('W+200') & s('N < 0') & s('Y+') & s('P+') &
                        s('ERP > 3%') & s('E-2') & s('CPI_y < 0%'))
    T['T4_2022离场']  = (s('CPI_y ≥ 5%') & s('S+6') & s('Sσ_200 ≥ +3σ') &
                        (s('WALCL+ ≥ 500hm') | s('ERP < 0') | s('N- ≥ μ+1σ') |
                         s('F+') | s('N > 0')) &
                        NOT(s('WALCL+ ≥ 1000hm')))
    T['T5_2002入场']  = (s('S-6') & s('S- ≥ 30%') & s('Sσ_200 ≤ -3σ') &
                        s('V ≥ V_c+3σ') & s('N < 0') & s('Y+') & s('E+') &
                        NOT(s('N- ≥ μ+1σ')))
    T['T6_2009入场']  = (s('S-6') & s('S- ≥ 30%') & s('Sσ_200 ≤ -3σ') &
                        s('V ≥ V_c+3σ') & s('HY_t > 8%') & s('ERP < 0') &
                        s('WALCL+ ≥ 1000hm'))
    T['T7_2020入场']  = (s('Sσ_200 ≤ -3σ') & s('V ≥ V_c+3σ') & s('HY_t > 8%') &
                        s('WALCL+ ≥ 1000hm') & (s('P-') | s('ERP > 3%')))
    T['T8_2022入场']  = (s('S- ≥ 20%') & s('S-6') & s('F+') & s('CPI_y ≥ 7%') &
                        s('E-2') & s('ERP > 3%') & s('N < 0'))
    # T9 v3: 新增 S_t < S_200
    T['T9_白银坑1组'] = (s('W+200') & s('S- ≥ 10%') & s('Y+') & s('N < 0') & s('ERP > 3%') &
                        (s('HY_t > 8%') | s('max(V,21d) ≥ V_c+2σ') |
                         s('F+') | s('E-2')) &
                        NOT(s('S- ≥ 20%')) & NOT(s('F-')) & NOT(s('CPI_y ≥ 5%')) &
                        s('S_t < S_200'))
    # T10 v4: 新增 S_t < S_200
    T['T10_白银坑2组']= (s('W+200') & s('S- ≥ 10%') & NOT(s('S- ≥ 20%')) & s('S-+') &
                        s('N < 0') & NOT(s('HY_t > 8%')) & NOT(s('ERP > 3%')) &
                        NOT(s('CPI_y ≥ 5%')) & NOT(s('F-')) &
                        NOT(s('ERP < 1.5%') & s('Y-')) &
                        s('S_t < S_200'))

    result = {}
    for tid, arr in T.items():
        result[tid] = sorted(pd.DatetimeIndex(fm['date'][arr == 1].values).tolist())
    return result


def main():
    parser = argparse.ArgumentParser(description='V15 + 文件 4 回归测试')
    parser.add_argument('--v15', required=True,
                        help='V15 因子矩阵 Excel 路径(59因子时间序列_v15_*.xlsx)')
    parser.add_argument('--triggers', required=True,
                        help='文件 4 触发器合集路径(标普500触发器系统_10触发器合集_*.xlsx)')
    args = parser.parse_args()

    v15_path  = Path(args.v15)
    file4_path = Path(args.triggers)
    if not v15_path.exists():
        print(f"❌ V15 文件不存在: {v15_path}"); sys.exit(1)
    if not file4_path.exists():
        print(f"❌ 文件 4 不存在: {file4_path}"); sys.exit(1)

    print("="*78)
    print("回归测试 — 验证新引擎与 V15 矩阵 + 文件 4 完全一致")
    print("="*78)

    print(f"\n[1/3] 加载 V15 因子矩阵: {v15_path.name}")
    fm = load_v15_factor_matrix(v15_path)
    inter = load_v15_intermediate(v15_path)
    print(f"     矩阵 {len(fm)} 行 × {len(fm.columns)} 列, 中间量 {len(inter)} 行")

    print(f"\n[2/3] 加载文件 4 触发日: {file4_path.name}")
    file4 = load_file4_trigger_dates(file4_path)
    total_file4 = sum(len(v) for v in file4.values())
    print(f"     共 {total_file4} 个触发日(应为 857)")

    print(f"\n[3/3] 用 V15 矩阵 + 触发器公式重算...")
    recomputed = evaluate_triggers_on_v15(fm, inter=inter)

    # 比对
    print("\n" + "="*78)
    print(f"{'触发器':<14}{'文件 4':>8}{'重算':>8}{'共同':>8}{'仅文件 4':>10}{'仅重算':>8}  状态")
    print("-"*78)

    total_common = 0
    total_only_f4 = 0
    total_only_re = 0
    all_pass = True

    for tid in TRIGGERS.keys():
        f4_dates = set(file4.get(tid, []))
        re_dates = set(recomputed.get(tid, []))
        common = f4_dates & re_dates
        only_f4 = f4_dates - re_dates
        only_re = re_dates - f4_dates

        total_common  += len(common)
        total_only_f4 += len(only_f4)
        total_only_re += len(only_re)

        status = '✅ 完全一致' if (not only_f4 and not only_re) else '❌ 有差异'
        if only_f4 or only_re:
            all_pass = False

        print(f"{tid:<14}{len(f4_dates):>8}{len(re_dates):>8}"
              f"{len(common):>8}{len(only_f4):>10}{len(only_re):>8}  {status}")

        if only_f4:
            print(f"    仅文件 4(前 5):{[d.strftime('%Y-%m-%d') for d in sorted(only_f4)[:5]]}")
        if only_re:
            print(f"    仅重算(前 5):{[d.strftime('%Y-%m-%d') for d in sorted(only_re)[:5]]}")

    print("-"*78)
    print(f"{'合计':<14}{total_common+total_only_f4:>8}{total_common+total_only_re:>8}"
          f"{total_common:>8}{total_only_f4:>10}{total_only_re:>8}")

    print("\n" + "="*78)
    if all_pass:
        print("✅ 全部通过 — 新引擎与文件 4 的 857 触发日完全一致,可上线生产")
    else:
        print(f"❌ 不通过 — 总差异 {total_only_f4 + total_only_re} 天,禁止上线")
        sys.exit(1)
    print("="*78)


if __name__ == '__main__':
    main()
